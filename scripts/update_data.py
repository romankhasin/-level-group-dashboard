#!/usr/bin/env python3
"""Refresh Level Group dashboard data for GitHub Pages."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import gzip
import html
import io
import json
import os
from pathlib import Path
import re
import time
import urllib.error
import urllib.parse
import urllib.request

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
METRIKA_HISTORY_PATH = DATA_DIR / "metrika.json"
TARGETADS_HISTORY_PATH = DATA_DIR / "targetads.json"
LATEST_PATH = DATA_DIR / "latest.json"
LATEST_SUMMARY_PATH = DATA_DIR / "latest-summary.json"
STATUS_PATH = DATA_DIR / "status.json"
GOOGLE_ARCHIVE_PATH = DATA_DIR / "google_archive.json"
JOURNAL_ARCHIVE_PATH = DATA_DIR / "journal_archive.json"
JOURNAL_PATH = ROOT / "journal.html"

START_DATE = dt.date(2026, 5, 1)
VOLGA_FACTS_START_DATE = dt.date(2026, 7, 1)
VOLGA_PROJECT_TOKEN = "lvol"
COUNTER_IDS = (53197618, 100470605)
METRIKA_QUALITY_CALL_GOAL_IDS = {
    53197618: 411053186,
    100470605: 411053614,
}
METRIKA_QUALITY_CALL_FIELD = "Первично-качественные звонки"
GOOGLE_WORKBOOK_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1chd_Rr_pZPZM83xfI2vx6gziNdLGRmdS/export?format=xlsx"
)
AVITO_GOOGLE_WORKBOOK_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "15lJoU_ylnP10SCyWpg6LV0zwXGV5SIxy/export?format=xlsx"
)
AUGUST_PRG_GOOGLE_WORKBOOK_URL = (
    "https://drive.usercontent.google.com/download?"
    "id=1zT8E9gAQSSXmW1e1T9-Wy1V9-C4esm3C&export=download"
)
# The supplied August media report is the authoritative source only for the
# two agreed PRG platforms.  Keep other programmatic platforms on their own
# sources instead of accidentally replacing their facts from this workbook.
AUGUST_PRG_PLATFORM_SOURCES = {"yandex", "mts"}
DEFAULT_TARGETADS_PROJECT_ID = 12787
TARGETADS_API_URL = "https://api.targetads.io"
TARGETADS_RAW_FIELDS = [
    "InteractionDate",
    "InteractionPlacementId",
    "InteractionCreativeId",
]
TARGETADS_POLL_INTERVAL_SECONDS = 5
TARGETADS_JOB_TIMEOUT_SECONDS = 20 * 60
GOOGLE_ACTUAL_COST_VAT_MULTIPLIER = 1.22


def request_json(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    body: dict | None = None,
    attempts: int = 4,
) -> dict:
    payload = None if body is None else json.dumps(body).encode("utf-8")
    request_headers = {"Accept": "application/json", **(headers or {})}
    if payload is not None:
        request_headers["Content-Type"] = "application/json"

    for attempt in range(1, attempts + 1):
        request = urllib.request.Request(
            url,
            data=payload,
            headers=request_headers,
            method="POST" if payload is not None else "GET",
        )
        try:
            with urllib.request.urlopen(request, timeout=180) as response:
                return json.load(response)
        except urllib.error.HTTPError as error:
            details = error.read().decode("utf-8", errors="replace")[:1000]
            if error.code not in {429, 500, 502, 503, 504} or attempt == attempts:
                raise RuntimeError(f"HTTP {error.code} for {url}: {details}") from error
        except (urllib.error.URLError, TimeoutError) as error:
            if attempt == attempts:
                raise RuntimeError(f"Request failed for {url}: {error}") from error
        time.sleep(min(2**attempt, 15))

    raise RuntimeError(f"Request failed for {url}")


def request_success(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    attempts: int = 4,
) -> None:
    """Make a request whose successful response may intentionally be empty."""
    request_headers = {"Accept": "application/json", **(headers or {})}
    for attempt in range(1, attempts + 1):
        request = urllib.request.Request(url, headers=request_headers)
        try:
            with urllib.request.urlopen(request, timeout=180) as response:
                response.read()
            return
        except urllib.error.HTTPError as error:
            details = error.read().decode("utf-8", errors="replace")[:1000]
            if error.code not in {429, 500, 502, 503, 504} or attempt == attempts:
                raise RuntimeError(f"HTTP {error.code} for {url}: {details}") from error
        except (urllib.error.URLError, TimeoutError) as error:
            if attempt == attempts:
                raise RuntimeError(f"Request failed for {url}: {error}") from error
        time.sleep(min(2**attempt, 15))

    raise RuntimeError(f"Request failed for {url}")


def download_file(url: str, destination: Path, attempts: int = 4) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(1, attempts + 1):
        try:
            request = urllib.request.Request(url, headers={"User-Agent": "LevelGroupDashboard/1.0"})
            with urllib.request.urlopen(request, timeout=240) as response:
                destination.write_bytes(response.read())
            return
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as error:
            if attempt == attempts:
                raise RuntimeError(f"Could not download {url}: {error}") from error
            time.sleep(min(2**attempt, 15))


def read_json_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        payload = payload.get("rows", [])
    return payload if isinstance(payload, list) else []


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def build_startup_summary(latest: dict) -> dict:
    """Return the tiny, safe-to-fetch payload used for the dashboard shell.

    Detailed Metrika and verifier rows intentionally stay in ``latest.json``.
    They are only downloaded after a person asks to open the detailed report.
    """
    raw_rows = latest.get("rawRows") or []
    verifier_rows = latest.get("verifierRows") or []
    return {
        "version": 1,
        "generatedAt": latest.get("generatedAt"),
        "period": latest.get("period") or {},
        "sourceFile": latest.get("sourceFile"),
        "status": {
            "targetads": {
                "enabled": bool((latest.get("status") or {}).get("targetads", {}).get("enabled")),
                "mode": (latest.get("status") or {}).get("targetads", {}).get("mode"),
            }
        },
        "counts": {
            "metrikaRows": len(raw_rows),
            "verifierRows": len(verifier_rows),
        },
        "detailUrl": "data/latest.json",
    }


def date_chunks(start: dt.date, end: dt.date, days: int = 31):
    cursor = start
    while cursor <= end:
        chunk_end = min(cursor + dt.timedelta(days=days - 1), end)
        yield cursor, chunk_end
        cursor = chunk_end + dt.timedelta(days=1)


def dimension_text(value: object) -> str:
    if isinstance(value, dict):
        return str(value.get("name") or value.get("id") or "").strip()
    return str(value or "").strip()


def fetch_metrika_period(
    token: str,
    counter_id: int,
    start: dt.date,
    end: dt.date,
) -> list[dict]:
    rows: list[dict] = []
    offset = 1
    limit = 100_000
    goal_id = METRIKA_QUALITY_CALL_GOAL_IDS[counter_id]

    while True:
        params = {
            "ids": str(counter_id),
            "date1": start.isoformat(),
            "date2": end.isoformat(),
            "dimensions": "ym:s:date,ym:s:lastsignUTMCampaign",
            "metrics": (
                "ym:s:visits,"
                "ym:s:bounceRate,"
                "ym:s:avgVisitDurationSeconds,"
                f"ym:s:goal{goal_id}reaches"
            ),
            "accuracy": "full",
            "limit": str(limit),
            "offset": str(offset),
        }
        url = "https://api-metrika.yandex.net/stat/v1/data?" + urllib.parse.urlencode(params)
        payload = request_json(url, headers={"Authorization": f"OAuth {token}"})
        page = payload.get("data") or []

        for item in page:
            dimensions = item.get("dimensions") or []
            metrics = item.get("metrics") or []
            if len(dimensions) < 2 or len(metrics) < 4:
                continue
            report_date = dimension_text(dimensions[0])
            campaign = dimension_text(dimensions[1])
            visits = int(round(float(metrics[0] or 0)))
            if not report_date or not campaign or visits <= 0:
                continue
            rows.append(
                {
                    "counter_id": counter_id,
                    "Дата визита": report_date,
                    "UTM Campaign": campaign,
                    "Визиты": visits,
                    "Отказы": float(metrics[1] or 0),
                    "Время на сайте": float(metrics[2] or 0),
                    METRIKA_QUALITY_CALL_FIELD: int(round(float(metrics[3] or 0))),
                }
            )

        total_rows = int(payload.get("total_rows") or len(page))
        if not page or offset - 1 + len(page) >= total_rows:
            break
        offset += len(page)

    return rows


def update_metrika(token: str, yesterday: dt.date) -> tuple[list[dict], dict]:
    existing = read_json_rows(METRIKA_HISTORY_PATH)
    keyed: dict[tuple[int, str, str], dict] = {}
    for row in existing:
        try:
            key = (
                int(row.get("counter_id") or 0),
                str(row.get("Дата визита") or ""),
                str(row.get("UTM Campaign") or ""),
            )
        except (TypeError, ValueError):
            continue
        if key[0] and key[1] and key[2]:
            keyed[key] = row

    fetched_count = 0
    ranges: dict[str, dict[str, str | int]] = {}
    for counter_id in COUNTER_IDS:
        needs_quality_calls_backfill = any(
            key[0] == counter_id
            and METRIKA_QUALITY_CALL_FIELD not in row
            for key, row in keyed.items()
        )
        counter_dates = [
            dt.date.fromisoformat(key[1])
            for key in keyed
            if key[0] == counter_id and key[1]
        ]
        if needs_quality_calls_backfill or not counter_dates:
            fetch_from = START_DATE
        else:
            fetch_from = max(counter_dates) + dt.timedelta(days=1)
        # Add the new Волга object from the beginning of July once.  The
        # regular incremental refresh cannot otherwise retrieve historical
        # July rows after the object mapping is introduced.
        has_volga_rows = any(
            key[0] == counter_id
            and str(row.get("UTM Campaign") or "").lower().startswith(VOLGA_PROJECT_TOKEN)
            for key, row in keyed.items()
        )
        if not has_volga_rows:
            fetch_from = min(fetch_from, VOLGA_FACTS_START_DATE)
        ranges[str(counter_id)] = {
            "from": fetch_from.isoformat(),
            "to": yesterday.isoformat(),
            "new_rows": 0,
            "quality_calls_backfill": needs_quality_calls_backfill,
        }
        if fetch_from > yesterday:
            continue

        for chunk_start, chunk_end in date_chunks(fetch_from, yesterday):
            new_rows = fetch_metrika_period(token, counter_id, chunk_start, chunk_end)
            for row in new_rows:
                key = (counter_id, row["Дата визита"], row["UTM Campaign"])
                keyed[key] = row
            fetched_count += len(new_rows)
            ranges[str(counter_id)]["new_rows"] += len(new_rows)

    rows = sorted(
        keyed.values(),
        key=lambda row: (
            str(row.get("Дата визита") or ""),
            int(row.get("counter_id") or 0),
            str(row.get("UTM Campaign") or ""),
        ),
    )
    write_json(METRIKA_HISTORY_PATH, {"rows": rows})
    return rows, {"new_rows": fetched_count, "total_rows": len(rows), "ranges": ranges}


def cell_date(value: object) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(text[:10], pattern).date()
        except ValueError:
            pass
    return None


def number(value: object) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def workbook_rows(sheet) -> tuple[list[str], list[tuple]]:
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(iterator)]
    return headers, list(iterator)


def read_google_workbook(path: Path, yesterday: dt.date) -> tuple[list[dict], list[dict], dict]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    required_sheets = {"Данные_метрика"}
    missing = required_sheets.difference(workbook.sheetnames)
    if missing:
        raise RuntimeError(f"Google workbook is missing sheets: {sorted(missing)}")

    metric_headers, metric_source_rows = workbook_rows(workbook["Данные_метрика"])
    metric_index = {name: index for index, name in enumerate(metric_headers)}
    required_metric_headers = {
        "report_date",
        "utm_source",
        "technical_mark",
        "impressions",
        "clicks",
        "cost",
    }
    missing_headers = required_metric_headers.difference(metric_index)
    if missing_headers:
        raise RuntimeError(f"Данные_метрика is missing columns: {sorted(missing_headers)}")

    aggregated: dict[tuple[str, str], dict] = {}
    for row in metric_source_rows:
        source = str(row[metric_index["utm_source"]] or "").strip().lower()
        if source not in {"yandex", "mts"}:
            continue
        report_date = cell_date(row[metric_index["report_date"]])
        technical_mark = str(row[metric_index["technical_mark"]] or "").strip()
        if not report_date or report_date < START_DATE or report_date > yesterday or not technical_mark:
            continue
        key = (report_date.isoformat(), technical_mark.lower())
        if key not in aggregated:
            aggregated[key] = {
                "placement_nm": technical_mark,
                "interaction_dt": report_date.isoformat(),
                "impressions": 0.0,
                "clicks": 0.0,
                "givt": 0.0,
                "fraud_impressions": 0.0,
                "cost": 0.0,
                "has_actual_cost": True,
                "metric_source": source,
            }
        item = aggregated[key]
        item["impressions"] += number(row[metric_index["impressions"]])
        item["clicks"] += number(row[metric_index["clicks"]])
        # The Yandex/MTS source reports actual spend without VAT, while all
        # dashboard plans already include VAT. Add 22% VAT only to actual cost.
        item["cost"] += number(row[metric_index["cost"]]) * GOOGLE_ACTUAL_COST_VAT_MULTIPLIER

    workbook.close()
    return (
        sorted(aggregated.values(), key=lambda row: (row["interaction_dt"], row["placement_nm"])),
        [],
        {
            "metric_rows": len(aggregated),
        },
    )


def read_journal_workbook(path: Path) -> tuple[list[dict], dict]:
    """Read the optimisation journal from the current reporting workbook."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Дневник оптимизации" not in workbook.sheetnames:
        raise RuntimeError("Journal workbook is missing sheet: Дневник оптимизации")

    journal_headers, journal_source_rows = workbook_rows(workbook["Дневник оптимизации"])
    expected_journal_headers = [
        "Проект",
        "Канал",
        "Площадка",
        "Период РК",
        "Период оптимизационных действий",
        "Комментарии по оптимизации",
        "Какие действия предпринимаем дальше",
    ]
    if journal_headers[:7] != expected_journal_headers:
        raise RuntimeError("Дневник оптимизации has an unexpected header layout")

    journal_rows = []
    seen_journal_rows: set[tuple[str, ...]] = set()
    for source_row in journal_source_rows:
        values = list(source_row[:7])
        if not any(value not in (None, "") for value in values):
            continue
        journal_row = {header: str(values[index] or "").strip() for index, header in enumerate(expected_journal_headers)}
        journal_key = tuple(journal_row[header] for header in expected_journal_headers)
        if journal_key not in seen_journal_rows:
            seen_journal_rows.add(journal_key)
            journal_rows.append(journal_row)

    platforms = sorted({row["Площадка"] for row in journal_rows if row["Площадка"]})
    workbook.close()
    return journal_rows, {"journal_rows": len(journal_rows), "journal_platforms": platforms}


def read_august_prg_workbook(path: Path, yesterday: dt.date) -> tuple[list[dict], dict]:
    """Read August PRG actuals from the supplied media workbook.

    This workbook is an Excel file in Drive and its tab is named either
    ``Данные метрика`` or ``Данные_метрика`` depending on the export version.
    Only the explicitly agreed August PRG platforms are imported.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    metric_sheet_name = next(
        (
            sheet_name
            for sheet_name in workbook.sheetnames
            if sheet_name.strip().lower().replace(" ", "_") == "данные_метрика"
        ),
        None,
    )
    if not metric_sheet_name:
        raise RuntimeError("August PRG workbook is missing sheet: Данные метрика")

    headers, source_rows = workbook_rows(workbook[metric_sheet_name])
    column_index = {name: index for index, name in enumerate(headers)}
    required_headers = {
        "report_date",
        "utm_source",
        "technical_mark",
        "impressions",
        "clicks",
        "cost",
    }
    missing_headers = required_headers.difference(column_index)
    if missing_headers:
        raise RuntimeError(
            "August PRG Данные метрика is missing columns: "
            f"{sorted(missing_headers)}"
        )

    aggregated: dict[tuple[str, str], dict] = {}
    source_rows_count = 0
    for row in source_rows:
        source = str(row[column_index["utm_source"]] or "").strip().lower()
        report_date = cell_date(row[column_index["report_date"]])
        technical_mark = str(row[column_index["technical_mark"]] or "").strip()
        campaign_tokens = [token for token in technical_mark.lower().split("_") if token]
        if (
            source not in AUGUST_PRG_PLATFORM_SOURCES
            or not report_date
            or report_date.year != 2026
            or report_date.month != 8
            or report_date > yesterday
            or not technical_mark
            or "prg" not in campaign_tokens
            or "aug26" not in campaign_tokens
        ):
            continue
        source_rows_count += 1
        key = (report_date.isoformat(), technical_mark.lower())
        if key not in aggregated:
            aggregated[key] = {
                "placement_nm": technical_mark,
                "interaction_dt": report_date.isoformat(),
                "impressions": 0.0,
                "clicks": 0.0,
                "givt": 0.0,
                "fraud_impressions": 0.0,
                "cost": 0.0,
                "has_actual_cost": True,
                "metric_source": "google_prg_august",
            }
        item = aggregated[key]
        item["impressions"] += number(row[column_index["impressions"]])
        item["clicks"] += number(row[column_index["clicks"]])
        item["cost"] += number(row[column_index["cost"]]) * GOOGLE_ACTUAL_COST_VAT_MULTIPLIER

    workbook.close()
    rows = sorted(aggregated.values(), key=lambda row: (row["interaction_dt"], row["placement_nm"]))
    return rows, {
        "sheet": metric_sheet_name,
        "source_rows": source_rows_count,
        "metric_rows": len(rows),
        "platforms": sorted(AUGUST_PRG_PLATFORM_SOURCES),
    }


def is_avito_med_mrk_campaign(campaign: str) -> bool:
    tokens = [token for token in campaign.strip().lower().split("_") if token]
    return "avito-ads" in tokens and any(channel in tokens for channel in ("med", "mrk"))


def read_august_avito_workbook(path: Path, yesterday: dt.date) -> tuple[list[dict], dict]:
    """Read non-zero August Avito MED/MRK facts from the supplied workbook.

    The workbook currently has Avito technical marks in ``Данные_метрика``.
    A row is emitted only once it contains a media fact, so a blank Avito row
    cannot overwrite the working Avito report used as a temporary fallback.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Данные_метрика" not in workbook.sheetnames:
        raise RuntimeError("August media workbook is missing sheet: Данные_метрика")

    headers, source_rows = workbook_rows(workbook["Данные_метрика"])
    column_index = {name: index for index, name in enumerate(headers)}
    required_headers = {"report_date", "utm_source", "technical_mark", "impressions", "clicks", "cost"}
    missing_headers = required_headers.difference(column_index)
    if missing_headers:
        raise RuntimeError(
            "August Avito Данные_метрика is missing columns: "
            f"{sorted(missing_headers)}"
        )

    aggregated: dict[tuple[str, str], dict] = {}
    source_rows_count = 0
    media_rows_count = 0
    for row in source_rows:
        source = str(row[column_index["utm_source"]] or "").strip().lower()
        report_date = cell_date(row[column_index["report_date"]])
        campaign = str(row[column_index["technical_mark"]] or "").strip()
        if (
            source != "avito-ads"
            or not report_date
            or report_date.year != 2026
            or report_date.month != 8
            or report_date > yesterday
            or not is_avito_med_mrk_campaign(campaign)
        ):
            continue
        source_rows_count += 1
        impressions = number(row[column_index["impressions"]])
        clicks = number(row[column_index["clicks"]])
        cost = number(row[column_index["cost"]])
        if not any((impressions, clicks, cost)):
            continue
        media_rows_count += 1
        key = (report_date.isoformat(), campaign.lower())
        if key not in aggregated:
            aggregated[key] = {
                "placement_nm": campaign,
                "interaction_dt": report_date.isoformat(),
                "impressions": 0.0,
                "clicks": 0.0,
                "givt": 0.0,
                "fraud_impressions": 0.0,
                "cost": 0.0,
                "has_actual_cost": True,
                "metric_source": "google_avito",
            }
        item = aggregated[key]
        item["impressions"] += impressions
        item["clicks"] += clicks
        # Like the existing Avito report, this field is net of VAT.
        item["cost"] += cost * GOOGLE_ACTUAL_COST_VAT_MULTIPLIER

    workbook.close()
    rows = sorted(aggregated.values(), key=lambda row: (row["interaction_dt"], row["placement_nm"]))
    return rows, {
        "sheet": "Данные_метрика",
        "source_rows": source_rows_count,
        "media_rows": media_rows_count,
        "metric_rows": len(rows),
    }


def read_avito_workbook(path: Path, yesterday: dt.date) -> tuple[list[dict], dict]:
    """Read Avito MED/MRK actuals from the shared media workbook's Data sheet."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Данные" not in workbook.sheetnames:
        raise RuntimeError("Avito workbook is missing sheet: Данные")

    headers, source_rows = workbook_rows(workbook["Данные"])
    column_index = {name: index for index, name in enumerate(headers)}
    required_headers = {"Дата", "utm_campaign", "Показы", "Клики", "Бюджет без НДС"}
    missing_headers = required_headers.difference(column_index)
    if missing_headers:
        raise RuntimeError(f"Avito Данные is missing columns: {sorted(missing_headers)}")

    aggregated: dict[tuple[str, str], dict] = {}
    source_rows_count = 0
    for row in source_rows:
        report_date = cell_date(row[column_index["Дата"]])
        campaign = str(row[column_index["utm_campaign"]] or "").strip()
        if (
            not report_date
            or report_date < START_DATE
            or report_date > yesterday
            or not is_avito_med_mrk_campaign(campaign)
        ):
            continue
        source_rows_count += 1
        key = (report_date.isoformat(), campaign.lower())
        if key not in aggregated:
            aggregated[key] = {
                "placement_nm": campaign,
                "interaction_dt": report_date.isoformat(),
                "impressions": 0.0,
                "clicks": 0.0,
                "givt": 0.0,
                "fraud_impressions": 0.0,
                "cost": 0.0,
                "has_actual_cost": True,
                "metric_source": "google_avito",
            }
        item = aggregated[key]
        item["impressions"] += number(row[column_index["Показы"]])
        item["clicks"] += number(row[column_index["Клики"]])
        item["cost"] += number(row[column_index["Бюджет без НДС"]]) * GOOGLE_ACTUAL_COST_VAT_MULTIPLIER

    workbook.close()
    rows = sorted(aggregated.values(), key=lambda row: (row["interaction_dt"], row["placement_nm"]))
    return rows, {"source_rows": source_rows_count, "metric_rows": len(rows)}


def targetads_project_id() -> int:
    value = os.environ.get("TARGETADS_PROJECT_ID", "").strip() or str(DEFAULT_TARGETADS_PROJECT_ID)
    try:
        project_id = int(value)
    except ValueError as error:
        raise RuntimeError("TARGETADS_PROJECT_ID must be a positive integer") from error
    if project_id <= 0:
        raise RuntimeError("TARGETADS_PROJECT_ID must be a positive integer")
    return project_id


def targetads_url(path: str, project_id: int) -> str:
    return f"{TARGETADS_API_URL}{path}?" + urllib.parse.urlencode({"project_id": project_id})


def ensure_targetads_response(payload: object, action: str) -> dict:
    if not isinstance(payload, dict):
        raise RuntimeError(f"Unexpected Target Ads response while {action}: not a JSON object")
    if payload.get("ErrorCode"):
        message = str(payload.get("ErrorMessage") or "no error message")
        raise RuntimeError(f"Target Ads API error while {action}: {payload['ErrorCode']}: {message}")
    return payload


def validate_targetads_token(token: str, project_id: int) -> dict:
    # The documented successful response is HTTP 200 and may have no body.
    request_success(
        targetads_url("/v1/meta/token_validate", project_id),
        headers={"Authorization": f"Bearer {token}"},
    )
    return {"valid": True}


def fetch_targetads_metadata(token: str, project_id: int) -> tuple[dict[str, str], dict[str, str]]:
    payload = ensure_targetads_response(
        request_json(
            targetads_url("/v1/meta/campaigns", project_id)
            + "&"
            + urllib.parse.urlencode({"active": "false", "include_creative": "true"}),
            headers={"Authorization": f"Bearer {token}"},
        ),
        "loading placement metadata",
    )
    metadata = payload.get("meta")
    if not isinstance(metadata, list):
        raise RuntimeError(
            "Unexpected Target Ads placement metadata response: "
            f"keys={sorted(payload.keys())}"
        )
    placements: dict[str, str] = {}
    creatives: dict[str, str] = {}
    for item in metadata:
        if not isinstance(item, dict):
            continue
        placement_id = str(item.get("placement_id") or "").strip()
        placement_name = str(item.get("placement_name") or "").strip()
        if placement_id:
            placements[placement_id] = placement_name or f"Placement {placement_id}"
        for creative in item.get("creatives") or []:
            if not isinstance(creative, dict):
                continue
            creative_id = str(creative.get("creative_id") or "").strip()
            creative_name = str(creative.get("creative_name") or "").strip()
            if creative_id:
                creatives[creative_id] = creative_name or f"Creative {creative_id}"
    return placements, creatives


def create_targetads_raw_job(
    token: str,
    project_id: int,
    interaction_type: str,
    start: dt.date,
    end: dt.date,
) -> str:
    payload = ensure_targetads_response(
        request_json(
            targetads_url("/v2/reports/raw_reports", project_id),
            headers={"Authorization": f"Bearer {token}"},
            body={
                "Fields": TARGETADS_RAW_FIELDS,
                "DateFrom": start.isoformat(),
                "DateTo": end.isoformat(),
                "InteractionType": interaction_type,
            },
        ),
        f"creating {interaction_type} raw-data job for {start}..{end}",
    )
    job_id = str(payload.get("job_id") or "").strip()
    if not job_id:
        raise RuntimeError(
            "Target Ads did not return job_id while creating raw-data job: "
            f"keys={sorted(payload.keys())}"
        )
    return job_id


def wait_for_targetads_raw_job(token: str, project_id: int, job_id: str) -> dict:
    deadline = time.monotonic() + TARGETADS_JOB_TIMEOUT_SECONDS
    while True:
        payload = ensure_targetads_response(
            request_json(
                targetads_url(f"/v2/jobs/{urllib.parse.quote(job_id, safe='')}", project_id),
                headers={"Authorization": f"Bearer {token}"},
            ),
            f"checking raw-data job {job_id}",
        )
        status = str(payload.get("status") or "").upper()
        if status == "DONE":
            if not payload.get("download_url"):
                raise RuntimeError(f"Target Ads job {job_id} is DONE without download_url")
            return payload
        if status in {"FAILED", "CANCELLED", "EXPIRED"}:
            message = str(payload.get("error_message") or "no error message")
            raise RuntimeError(f"Target Ads job {job_id} ended as {status}: {message}")
        if status not in {"CREATED", "PROCESSING"}:
            raise RuntimeError(f"Target Ads job {job_id} returned unknown status {status!r}")
        if time.monotonic() >= deadline:
            raise RuntimeError(f"Timed out waiting for Target Ads job {job_id}")
        time.sleep(TARGETADS_POLL_INTERVAL_SECONDS)


def aggregate_targetads_csv(download_url: str, interaction_type: str) -> dict[tuple[str, str, str], int]:
    """Count a gzip-compressed Raw Data CSV without storing individual events."""
    aggregated: dict[tuple[str, str, str], int] = {}
    request = urllib.request.Request(download_url, headers={"User-Agent": "LevelGroupDashboard/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=300) as response:
            with gzip.GzipFile(fileobj=response) as compressed:
                with io.TextIOWrapper(compressed, encoding="utf-8-sig", newline="") as text_stream:
                    reader = csv.DictReader(text_stream)
                    required = {
                        "InteractionDate",
                        "InteractionPlacementId",
                        "InteractionCreativeId",
                    }
                    if not reader.fieldnames or not required.issubset(reader.fieldnames):
                        raise RuntimeError(
                            "Target Ads raw CSV has unexpected columns: "
                            f"{reader.fieldnames or []}"
                        )
                    for item in reader:
                        report_date = str(item.get("InteractionDate") or "").strip()
                        placement_id = str(item.get("InteractionPlacementId") or "").strip()
                        creative_id = str(item.get("InteractionCreativeId") or "").strip()
                        if cell_date(report_date) and placement_id and creative_id:
                            key = (report_date, creative_id, placement_id)
                            aggregated[key] = aggregated.get(key, 0) + 1
    except (OSError, UnicodeError, urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as error:
        raise RuntimeError(f"Could not download or read Target Ads {interaction_type} raw CSV: {error}") from error
    return aggregated


def fetch_targetads_period(
    token: str,
    project_id: int,
    placements: dict[str, str],
    creatives: dict[str, str],
    start: dt.date,
    end: dt.date,
) -> tuple[list[dict], dict]:
    metrics: dict[tuple[str, str], dict] = {}
    jobs: list[dict[str, object]] = []
    for interaction_type, metric in (("Impression", "impressions"), ("Click", "clicks")):
        job_id = create_targetads_raw_job(token, project_id, interaction_type, start, end)
        job = wait_for_targetads_raw_job(token, project_id, job_id)
        counts = aggregate_targetads_csv(str(job["download_url"]), interaction_type)
        jobs.append(
            {
                "id": job_id,
                "type": interaction_type,
                "status": "DONE",
                "api_row_count": int(job.get("row_count") or 0),
                "aggregated_events": sum(counts.values()),
            }
        )
        for (report_date, creative_id, placement_id), count in counts.items():
            values = metrics.setdefault(
                (report_date, creative_id),
                {"impressions": 0, "clicks": 0, "placement_ids": set()},
            )
            values[metric] += count
            values["placement_ids"].add(placement_id)

    rows = []
    for (report_date, creative_id), values in metrics.items():
        placement_ids = sorted(values["placement_ids"])
        fallback_placement = placements.get(placement_ids[0], "") if placement_ids else ""
        rows.append(
            {
                "creative_id": creative_id,
                "placement_id": ",".join(placement_ids),
                # Creative names are the dashboard's technical campaign keys.
                "placement_nm": creatives.get(
                    creative_id,
                    fallback_placement or f"Creative {creative_id}",
                ),
                "interaction_dt": report_date,
                "impressions": values["impressions"],
                "clicks": values["clicks"],
                # Raw Data API v2 does not provide GIVT or SIVT. Keep the
                # existing dashboard schema until Fraud Reports API is added.
                "givt": 0.0,
                "fraud_impressions": 0.0,
                "cost": 0.0,
                "has_actual_cost": False,
                "metric_source": "targetads_raw_v2",
            }
        )
    return rows, {"from": start.isoformat(), "to": end.isoformat(), "jobs": jobs}


def targetads_row_key(row: dict) -> tuple[str, str]:
    report_date = str(row.get("interaction_dt") or "")
    creative_id = str(row.get("creative_id") or "").strip()
    placement_id = str(row.get("placement_id") or "").strip()
    placement_name = str(row.get("placement_nm") or "").strip().lower()
    return report_date, creative_id or placement_id or placement_name


def update_targetads(token: str, yesterday: dt.date) -> tuple[list[dict], dict]:
    project_id = targetads_project_id()
    validate_targetads_token(token, project_id)
    placements, creatives = fetch_targetads_metadata(token, project_id)
    existing = read_json_rows(TARGETADS_HISTORY_PATH)
    needs_creative_backfill = any(
        not str(row.get("creative_id") or "").strip() for row in existing
    )
    keyed = {
        targetads_row_key(row): row
        for row in existing
        if row.get("interaction_dt") and row.get("placement_nm")
    }
    if needs_creative_backfill:
        keyed = {}
    existing_dates = [dt.date.fromisoformat(key[0]) for key in keyed if key[0]]
    api_min_date = yesterday - dt.timedelta(days=89)
    desired_fetch_from = (
        START_DATE
        if needs_creative_backfill
        else max(existing_dates) + dt.timedelta(days=1) if existing_dates else START_DATE
    )
    fetch_from = max(desired_fetch_from, api_min_date)
    fetched_count = 0
    completed_jobs: list[dict] = []
    if fetch_from <= yesterday:
        for chunk_start, chunk_end in date_chunks(fetch_from, yesterday, days=3):
            new_rows, period_status = fetch_targetads_period(
                token, project_id, placements, creatives, chunk_start, chunk_end
            )
            for row in new_rows:
                keyed[targetads_row_key(row)] = row
            fetched_count += len(new_rows)
            completed_jobs.extend(period_status["jobs"])

    rows = sorted(keyed.values(), key=lambda row: (row["interaction_dt"], row["placement_nm"]))
    write_json(TARGETADS_HISTORY_PATH, {"rows": rows})
    warning = None
    if not placements:
        warning = "Token is valid but Meta API returned no accessible placements"
    elif not creatives:
        warning = "Meta API returned no creative names; dashboard keys will use placement names"
    elif desired_fetch_from < api_min_date:
        warning = (
            "Raw Data API v2 only retains 90 days; "
            f"the requested history before {api_min_date.isoformat()} could not be loaded"
        )
    return rows, {
        "enabled": True,
        "configured": True,
        "api": "raw_data_v2",
        "project_id": project_id,
        "token_valid": True,
        "campaigns_available": len(placements),
        "creatives_available": len(creatives),
        "identity_field": "creative_name",
        "creative_backfill": needs_creative_backfill,
        "new_rows": fetched_count,
        "total_rows": len(rows),
        "from": fetch_from.isoformat(),
        "to": yesterday.isoformat(),
        "jobs_completed": len(completed_jobs),
        "events_aggregated": sum(
            int(job["aggregated_events"]) for job in completed_jobs
        ),
        "job_details": completed_jobs,
        "warning": warning,
    }


def is_yandex_mts_prg_campaign(campaign: str) -> bool:
    tokens = [token for token in campaign.strip().lower().split("_") if token]
    return "prg" in tokens and any(
        platform in tokens for platform in ("yandex", "mts")
    )


def is_google_august_prg_row(row: dict) -> bool:
    """August PRG media facts supplied by the dedicated Google workbook."""
    return (
        str(row.get("metric_source") or "").strip().lower() == "google_prg_august"
        and "prg" in [token for token in str(row.get("placement_nm") or "").lower().split("_") if token]
    )


def is_google_avito_med_mrk_row(row: dict) -> bool:
    return (
        str(row.get("metric_source") or "").strip().lower() == "google_avito"
        and is_avito_med_mrk_campaign(str(row.get("placement_nm") or ""))
    )


def google_row_has_media_facts(row: dict) -> bool:
    return any(number(row.get(metric)) != 0 for metric in ("impressions", "clicks", "cost"))


def combine_google_media_with_targetads_fraud(google_row: dict, targetads_row: dict) -> dict:
    """Keep Google media facts and attach fraud facts reported by Target Ads."""
    combined = dict(google_row)
    combined["givt"] = number(targetads_row.get("givt"))
    combined["fraud_impressions"] = number(targetads_row.get("fraud_impressions"))
    return combined


def verifier_campaign_key(campaign: str) -> str:
    """Normalize harmless extra separators before joining source rows."""
    return "_".join(token for token in campaign.strip().lower().split("_") if token)


def combine_targetads_rows(existing: dict, incoming: dict) -> dict:
    """Sum Target Ads rows that share a reporting date and campaign key."""
    combined = dict(existing)
    for metric in ("impressions", "clicks", "givt", "fraud_impressions", "cost"):
        combined[metric] = number(existing.get(metric)) + number(incoming.get(metric))
    combined["has_actual_cost"] = bool(
        existing.get("has_actual_cost") or incoming.get("has_actual_cost")
    )
    return combined


def merge_verifier_rows(targetads_rows: list[dict], google_rows: list[dict]) -> list[dict]:
    """Merge media and fraud facts without duplicating Google and Target Ads.

    Google has priority for Yandex/MTS PRG, the August Yandex/MTS PRG facts,
    and Avito MED/MRK only when it has at least one non-zero
    media fact. Otherwise Target Ads supplies media.
    Target Ads rows with the same date and campaign are summed before sources
    are merged, preventing creative-level rows from being overwritten.
    """
    merged: dict[tuple[str, str], dict] = {}
    for source_row in targetads_rows:
        row = dict(source_row)
        campaign = str(row.get("placement_nm") or "")
        report_date = str(row.get("interaction_dt") or "")
        if not report_date or not campaign:
            continue
        key = (report_date, verifier_campaign_key(campaign))
        merged[key] = combine_targetads_rows(merged[key], row) if key in merged else row

    for row in google_rows:
        report_date = str(row.get("interaction_dt") or "")
        campaign = str(row.get("placement_nm") or "")
        if not report_date or not campaign:
            continue
        key = (report_date, verifier_campaign_key(campaign))
        targetads_row = merged.get(key)
        google_owned = (
            is_yandex_mts_prg_campaign(campaign)
            or is_google_august_prg_row(row)
            or is_google_avito_med_mrk_row(row)
        )
        if google_owned and google_row_has_media_facts(row):
            merged[key] = (
                combine_google_media_with_targetads_fraud(row, targetads_row)
                if targetads_row
                else row
            )
        elif key not in merged:
            merged[key] = row

    return sorted(merged.values(), key=lambda row: (row["interaction_dt"], row["placement_nm"]))


def merge_google_rows(archive_rows: list[dict], live_rows: list[dict]) -> list[dict]:
    """Combine the fixed historical archive with the current Google report.

    Live rows win if a date + technical_mark pair ever exists in both sources.
    """
    merged: dict[tuple[str, str], dict] = {}
    for row in (*archive_rows, *live_rows):
        report_date = str(row.get("interaction_dt") or "").strip()
        technical_mark = str(row.get("placement_nm") or "").strip()
        if not report_date or not technical_mark:
            continue
        merged[(report_date, technical_mark.lower())] = row
    return sorted(merged.values(), key=lambda row: (row["interaction_dt"], row["placement_nm"]))


def journal_html(rows: list[dict], generated_at: str) -> str:
    def text(value: object, fallback: str = "") -> str:
        value_text = str(value or "").strip()
        return value_text or fallback

    def period_label(value: str) -> str:
        try:
            return dt.datetime.strptime(value[:19], "%Y-%m-%d %H:%M:%S").strftime("%d.%m.%Y")
        except ValueError:
            return value

    def period_sort_key(value: str) -> dt.date:
        try:
            return dt.datetime.strptime(value[:10], "%Y-%m-%d").date()
        except ValueError:
            matches = re.findall(r"(\d{1,2})\.(\d{1,2})", value)
            if matches:
                day, month = matches[-1]
                try:
                    return dt.date(2026, int(month), int(day))
                except ValueError:
                    pass
        return dt.date.min

    period_groups: dict[str, list[dict]] = {}
    for row in rows:
        period = text(row.get("Период оптимизационных действий"), "Без даты")
        period_groups.setdefault(period, []).append(row)

    period_blocks = []
    sorted_periods = sorted(period_groups, key=period_sort_key, reverse=True)
    for period_index, period in enumerate(sorted_periods):
        period_rows = period_groups[period]
        platform_groups: dict[str, list[dict]] = {}
        for row in period_rows:
            platform = text(row.get("Площадка"), "Площадка не указана")
            platform_groups.setdefault(platform, []).append(row)

        platform_blocks = []
        for platform in sorted(platform_groups, key=str.casefold):
            cards = []
            for row in platform_groups[platform]:
                project = html.escape(text(row.get("Проект"), "Без проекта"))
                channel = html.escape(text(row.get("Канал")))
                campaign_period = html.escape(text(row.get("Период РК")))
                comment = html.escape(text(row.get("Комментарии по оптимизации"))).replace("\n", "<br>")
                next_steps = html.escape(text(row.get("Какие действия предпринимаем дальше"))).replace("\n", "<br>")
                meta = " · ".join(value for value in (channel, campaign_period) if value)
                next_block = (
                    f'<div class="next"><strong>Следующие действия</strong><div>{next_steps}</div></div>'
                    if next_steps
                    else ""
                )
                cards.append(
                    f"""
                    <article class="card">
                      <div class="card-head"><h3>{project}</h3><span>{meta}</span></div>
                      <div class="comment">{comment or 'Комментарий не заполнен'}</div>
                      {next_block}
                    </article>
                    """
                )

            platform_blocks.append(
                f"""
                <details class="platform-group">
                  <summary><span>{html.escape(platform)}</span><small>{len(cards)} записей</small></summary>
                  <div class="cards">{''.join(cards)}</div>
                </details>
                """
            )

        open_attr = " open" if period_index == 0 else ""
        period_blocks.append(
            f"""
            <details class="period-group"{open_attr}>
              <summary><span>{html.escape(period_label(period))}</span><small>{len(platform_groups)} площадок · {len(period_rows)} записей</small></summary>
              <div class="platforms">{''.join(platform_blocks)}</div>
            </details>
            """
        )

    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Журнал оптимизаций — Level Group</title>
  <style>
    :root{{--ink:#20233a;--muted:#747782;--line:#e9e4da;--blue:#ddeef8;--paper:#f7f4ef;--green:#2f7d5a}}
    *{{box-sizing:border-box}} body{{margin:0;font-family:Arial,sans-serif;color:var(--ink);background:linear-gradient(145deg,#fff,var(--paper))}}
    header{{padding:32px 20px;background:linear-gradient(135deg,#fff,var(--blue));border-bottom:1px solid var(--line)}}
    header>div,main{{max-width:1180px;margin:auto}} h1{{margin:0 0 8px;font-size:clamp(28px,5vw,52px)}}
    header p{{margin:0;color:var(--muted);line-height:1.6}} .back{{display:inline-block;margin-top:18px;color:var(--ink);font-weight:700}}
    main{{padding:28px 20px 60px;display:grid;gap:18px}} summary{{cursor:pointer;list-style:none}} summary::-webkit-details-marker{{display:none}}
    .period-group{{background:#fff;border:1px solid var(--line);border-radius:24px;overflow:hidden;box-shadow:0 12px 34px rgba(32,35,58,.06)}}
    .period-group>summary,.platform-group>summary{{display:flex;justify-content:space-between;align-items:center;gap:16px;font-weight:800}}
    .period-group>summary{{padding:22px 24px;font-size:22px;background:linear-gradient(135deg,#fff,var(--blue))}}
    summary small{{color:var(--muted);font-size:13px;font-weight:600}} summary span::before{{content:'›';display:inline-block;margin-right:12px;transition:transform .2s}}
    details[open]>summary span::before{{transform:rotate(90deg)}} .platforms{{padding:14px;display:grid;gap:12px}}
    .platform-group{{border:1px solid var(--line);border-radius:18px;overflow:hidden;background:#fbfaf8}} .platform-group>summary{{padding:17px 18px;font-size:17px}}
    .cards{{padding:0 12px 12px;display:grid;gap:12px}} .card{{background:#fff;border:1px solid var(--line);border-radius:16px;padding:18px}}
    .card-head{{display:flex;justify-content:space-between;gap:18px;align-items:flex-start;border-bottom:1px solid var(--line);padding-bottom:14px;margin-bottom:16px}}
    h3{{margin:0;font-size:18px}} .card-head span{{color:var(--muted);font-size:13px;text-align:right}}
    .comment,.next{{line-height:1.65;font-size:14px}} .next{{margin-top:16px;padding:16px;border-radius:16px;background:#f2f8f4}} .next strong{{display:block;margin-bottom:8px}}
    @media(max-width:700px){{.card-head{{display:block}}.card-head span{{display:block;text-align:left;margin-top:8px}}}}
  </style>
</head>
<body>
  <header><div><h1>Журнал оптимизаций</h1><p>Все площадки · обновлено {html.escape(generated_at)}</p><a class="back" href="./">← Вернуться в дашборд</a></div></header>
  <main>{''.join(period_blocks) if period_blocks else '<p>Записи пока отсутствуют.</p>'}</main>
</body>
</html>
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-metrika", action="store_true", help="Local validation only")
    parser.add_argument("--workbook", type=Path, help="Use a local workbook instead of downloading")
    args = parser.parse_args()

    now_utc = dt.datetime.now(dt.timezone.utc)
    yesterday = now_utc.astimezone(dt.timezone(dt.timedelta(hours=3))).date() - dt.timedelta(days=1)
    if yesterday < START_DATE:
        raise RuntimeError("Yesterday is earlier than the configured report start date")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    token = os.environ.get("YANDEX_METRIKA_TOKEN", "").strip()
    if args.skip_metrika:
        metrika_rows = read_json_rows(METRIKA_HISTORY_PATH)
        metrika_status = {"skipped": True, "total_rows": len(metrika_rows)}
    else:
        if not token:
            raise RuntimeError("YANDEX_METRIKA_TOKEN is not configured")
        metrika_rows, metrika_status = update_metrika(token, yesterday)

    workbook_path = args.workbook or (ROOT / ".cache" / "level_group_report.xlsx")
    if args.workbook is None:
        download_file(GOOGLE_WORKBOOK_URL, workbook_path)
    live_google_rows, _, google_status = read_google_workbook(workbook_path, yesterday)
    avito_workbook_path = ROOT / ".cache" / "avito_media_report.xlsx"
    download_file(AVITO_GOOGLE_WORKBOOK_URL, avito_workbook_path)
    avito_google_rows, avito_status = read_avito_workbook(avito_workbook_path, yesterday)
    august_prg_workbook_path = ROOT / ".cache" / "august_prg_media_report.xlsx"
    download_file(AUGUST_PRG_GOOGLE_WORKBOOK_URL, august_prg_workbook_path)
    august_prg_google_rows, august_prg_status = read_august_prg_workbook(
        august_prg_workbook_path, yesterday
    )
    journal_rows, journal_status = read_journal_workbook(august_prg_workbook_path)
    # Keep verified historical cards that are not present in the live source.
    journal_rows = [*read_json_rows(JOURNAL_ARCHIVE_PATH), *journal_rows]
    august_avito_google_rows, august_avito_status = read_august_avito_workbook(
        august_prg_workbook_path, yesterday
    )
    archived_google_rows = read_json_rows(GOOGLE_ARCHIVE_PATH)
    google_rows = merge_google_rows(
        archived_google_rows,
        # Facts from the supplied August workbook are last so they replace the
        # temporary Avito fallback only for matching date + technical mark.
        [
            *live_google_rows,
            *avito_google_rows,
            *august_prg_google_rows,
            *august_avito_google_rows,
        ],
    )
    google_status.update(
        {
            "archive_rows": len(archived_google_rows),
            "live_rows": len(live_google_rows),
            "avito": avito_status,
            "august_avito": august_avito_status,
            "august_prg": august_prg_status,
            "metric_rows": len(google_rows),
        }
    )

    # Target Ads is deliberately manual-only for now. Do not invoke Raw Data
    # API v2 and do not carry the last automatic API snapshot into latest.json:
    # the dashboard merges a manually uploaded Target Ads file in the browser.
    targetads_rows: list[dict] = []
    targetads_status = {
        "enabled": False,
        "mode": "manual_upload_only",
        "message": "Target Ads API update is disabled; upload the Target Ads report manually in the dashboard",
        "total_rows": 0,
    }

    verifier_rows = merge_verifier_rows(targetads_rows, google_rows)
    generated_at = now_utc.isoformat().replace("+00:00", "Z")
    latest = {
        "version": 1,
        "generatedAt": generated_at,
        "period": {"from": START_DATE.isoformat(), "to": yesterday.isoformat()},
        "rawRows": metrika_rows,
        "verifierRows": verifier_rows,
        "sourceFile": "Автоматическая выгрузка Яндекс Метрики",
        "verifierFile": (
            "Google Данные_метрика + августовский PRG Google-отчёт "
            "+ Avito Данные + фиксированный архив июня"
        ),
        "status": {
            "metrika": metrika_status,
            "google": google_status,
            "journal": journal_status,
            "targetads": targetads_status,
        },
    }
    write_json(LATEST_PATH, latest)
    write_json(LATEST_SUMMARY_PATH, build_startup_summary(latest))
    write_json(STATUS_PATH, {"generatedAt": generated_at, **latest["status"]})
    JOURNAL_PATH.write_text(journal_html(journal_rows, generated_at), encoding="utf-8")

    print(
        json.dumps(
            {
                "generatedAt": generated_at,
                "metrikaRows": len(metrika_rows),
                "verifierRows": len(verifier_rows),
                "journalRows": len(journal_rows),
                "targetAdsEnabled": False,
                "targetAdsMode": "manual_upload_only",
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
